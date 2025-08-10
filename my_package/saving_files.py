from openpyxl import load_workbook
from openpyxl import Workbook
import pandas as pd

def save_cycle_data_to_common_excel(df_cycle_data, excel_file, new_sheet_name):
    ''' This function saves cycle data from different batteries to an excel file 
    so that the batteries ageing curves can be compared afterwards
    '''

    # Try to load the existing Excel file, create a new one if it doesn't exist
    try:
        workbook = load_workbook(excel_file)
    except FileNotFoundError:
        workbook = Workbook()
        workbook.save(excel_file)
        workbook = load_workbook(excel_file)

    # Remove the sheet if it already exists
    if new_sheet_name in workbook.sheetnames:
        del workbook[new_sheet_name]
        workbook.save(excel_file)

    # Write the DataFrame to the new sheet
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a') as writer:
        df_cycle_data.to_excel(writer, sheet_name=new_sheet_name, index=False)

    print(f'DataFrame saved in the sheet {new_sheet_name} of the Excel file.')